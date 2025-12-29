import threading
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from ci_data_cmd import consultar_cedula
import pandas as pd

class CedulaGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Consulta de Cédulas")
        self.root.geometry("1150x600")

        self.row_count = 0

        paned = ttk.PanedWindow(root, orient=tk.HORIZONTAL)
        paned.pack(fill="both", expand=True, padx=10, pady=10)

        left_frame = ttk.Frame(paned, padding=10)
        paned.add(left_frame, weight=1)

        ttk.Label(left_frame,
                  text="Ingrese cédulas (una por línea):",
                  font=("Segoe UI", 10, "bold")).pack(anchor="w")

        text_frame = ttk.Frame(left_frame)
        text_frame.pack(fill="both", expand=True)

        self.text_input = tk.Text(text_frame, height=25, width=35)
        self.text_input.pack(side="left", fill="both", expand=True)

        text_scroll = ttk.Scrollbar(
            text_frame,
            orient="vertical",
            command=self.text_input.yview
        )
        text_scroll.pack(side="right", fill="y")
        self.text_input.config(yscrollcommand=text_scroll.set)

        btn_frame = ttk.Frame(left_frame)
        btn_frame.pack(fill="x", pady=5)

        self.btn_nueva = ttk.Button(
            btn_frame,
            text="Nueva consulta",
            command=lambda: self.iniciar_consulta(clear=True)
        )
        self.btn_nueva.pack(side="left", padx=5)

        self.btn_seguir = ttk.Button(
            btn_frame,
            text="Seguir consulta",
            command=lambda: self.iniciar_consulta(clear=False)
        )
        self.btn_seguir.pack(side="left", padx=5)

        self.status_label = ttk.Label(left_frame, text="Listo", foreground="blue")
        self.status_label.pack(anchor="w", pady=5)

        right_frame = ttk.Frame(paned, padding=10)
        paned.add(right_frame, weight=3)

        ttk.Label(right_frame,
                  text="Resultados",
                  font=("Segoe UI", 10, "bold")).pack(anchor="w")

        table_frame = ttk.Frame(right_frame)
        table_frame.pack(fill="both", expand=True)

        columns = ("num", "cedula", "nombre", "fecha_larga", "fecha_corta", "edad")

        self.tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            height=20
        )

        self.tree.heading("num", text="N°")
        self.tree.heading("cedula", text="Cédula")
        self.tree.heading("nombre", text="Nombre Completo")
        self.tree.heading("fecha_larga", text="Fecha Nacimiento Larga")
        self.tree.heading("fecha_corta", text="Fecha Nacimiento Corta")
        self.tree.heading("edad", text="Edad")

        self.tree.column("num", width=50, anchor="center")
        self.tree.column("cedula", width=110)
        self.tree.column("nombre", width=220)
        self.tree.column("fecha_larga", width=150)
        self.tree.column("fecha_corta", width=150)
        self.tree.column("edad", width=80, anchor="center")

        self.tree.pack(side="left", fill="both", expand=True)

        table_scroll = ttk.Scrollbar(
            table_frame,
            orient="vertical",
            command=self.tree.yview
        )
        table_scroll.pack(side="right", fill="y")
        self.tree.config(yscrollcommand=table_scroll.set)

        ttk.Button(
            right_frame,
            text="Exportar a Excel",
            command=self.exportar_excel
        ).pack(pady=5)

        self.center_window()

    def center_window(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f"{width}x{height}+{x}+{y-50}")

    def iniciar_consulta(self, clear):
        thread = threading.Thread(target=self.consultar, args=(clear,))
        thread.start()

    def consultar(self, clear):
        if clear:
            self.tree.delete(*self.tree.get_children())
            self.row_count = 0

        cedulas = self.text_input.get("1.0", tk.END).strip().split("\n")
        cedulas = [c.strip() for c in cedulas if c.strip()]

        if not cedulas:
            messagebox.showwarning("Atención", "Debe ingresar al menos una cédula.")
            return

        self.status_label.config(text="Consultando...", foreground="orange")
        self.btn_nueva.config(state="disabled")
        self.btn_seguir.config(state="disabled")

        for ci in cedulas:
            self.row_count += 1

            if not ci.isdigit() or len(ci) != 10:
                self.tree.insert(
                    "",
                    "end",
                    values=(self.row_count, ci, "Cédula inválida", "", "", "")
                )
                continue

            c, n, f_texto, f_iso, edad = consultar_cedula(ci)

            if c:
                self.tree.insert(
                    "",
                    "end",
                    values=(self.row_count, c, n, f_texto, f_iso, edad)
                )
            else:
                self.tree.insert(
                    "",
                    "end",
                    values=(self.row_count, ci, "No encontrado", "", "", "")
                )

        self.status_label.config(text="Consulta finalizada", foreground="green")
        self.btn_nueva.config(state="normal")
        self.btn_seguir.config(state="normal")

    def exportar_excel(self):
        data = []
        for row in self.tree.get_children():
            data.append(self.tree.item(row)["values"])

        if not data:
            messagebox.showwarning("Atención", "No hay datos para exportar.")
            return

        df = pd.DataFrame(
            data,
            columns=[
                "N°",
                "Cédula",
                "Nombre",
                "Fecha Nacimiento Larga",
                "Fecha Nacimiento Corta",
                "Edad"
            ]
        )
        
        df["Fecha Nacimiento Corta"] = pd.to_datetime(
            df["Fecha Nacimiento Corta"],
            format="%Y-%m-%d",
            errors="coerce"
        )

        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Guardar archivo"
        )

        if archivo:
            from openpyxl import load_workbook

            with pd.ExcelWriter(archivo, engine="openpyxl") as writer:
                df.to_excel(writer, index=False)

            wb = load_workbook(archivo)
            ws = wb.active

            col = 5

            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col).number_format = "dd/mm/yyyy"

            wb.save(archivo)

            messagebox.showinfo("Éxito", "Archivo exportado correctamente.")

if __name__ == "__main__":
    root = tk.Tk()
    CedulaGUI(root)
    root.mainloop()
